#!/usr/bin/env python3
"""Build an executive dashboard workbook from the CSA Monthly Reminder export."""
import json
from collections import Counter, OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import PieChart, BarChart, Reference, DoughnutChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

SRC = "CSA_Monthly_reminder_VK_6222026716AM.xlsx"
OUT = "CSA_Monthly_Reminder_Executive_Dashboard.xlsx"

# ---------------- Load + derive ----------------
wb_in = openpyxl.load_workbook(SRC, data_only=True)
ws_in = wb_in["Archer Search Report"]
rows = list(ws_in.iter_rows(min_row=2, values_only=True))
header = [h for h in rows[0]][1:]
records = []
for r in rows[1:]:
    vals = list(r)[1:]
    if any(v is not None for v in vals):
        records.append(dict(zip(header, vals)))

def family(n):
    n = (n or "").lower()
    if "leaver" in n: return "Leaver Access Review"
    if "mover" in n: return "Mover Access Review"
    if any(k in n for k in ("firefighter","privileged","super","admin")): return "Privileged / Admin Access"
    if any(k in n for k in ("access request","creation and change","operational access")): return "Access Provisioning / Approval"
    if "periodically reviewed" in n or "access rights to" in n: return "Periodic Recertification"
    return "Other"

def cadence(r):
    r = (r or "").lower()
    if "monthly" in r: return "Monthly"
    if "quarterly" in r: return "Quarterly"
    if "annual" in r: return "Annually"
    return "Unspecified"

for d in records:
    d["_family"] = family(d.get("Control Procedure Name"))
    d["_cadence"] = cadence(d.get("Remarks"))
    d["_hasDele"] = bool(d.get("Control Performer Delegate"))
    d["_relaunch"] = any(k in (d.get("Remarks") or "").lower() for k in ("relaunch","to be","mail to be sent"))

total = len(records)
def cnt(f): return Counter(f(d) for d in records)
by_scope = cnt(lambda d: d.get("Risk/Control Scope") or "Blank")
by_family = cnt(lambda d: d["_family"])
by_cadence = cnt(lambda d: d["_cadence"])
by_perf = cnt(lambda d: (d.get("Control Performer") or "").rsplit(" ",1)[0])
critical = by_scope.get("Critical",0)
no_dele = sum(1 for d in records if not d["_hasDele"])
crit_no_dele = sum(1 for d in records if d.get("Risk/Control Scope")=="Critical" and not d["_hasDele"])
relaunch = sum(1 for d in records if d["_relaunch"])
performers = len(by_perf)
top_perf = by_perf.most_common(1)[0]
no_entity = sum(1 for d in records if not d.get("Entity Code"))
ag = sum(1 for d in records if d.get("Business Process")=="Access Governance")
itgc = sum(1 for d in records if d.get("Business Sub Process")=="ITGC")

# ---------------- Styling helpers ----------------
NAVY="1F2A44"; BLUE="2E5BFF"; SLATE="334155"; LIGHT="EEF2FB"; WHITE="FFFFFF"
RED="E23D3D"; AMBER="E8930C"; GREEN="2BA84A"; GREY="64748B"
SCOPE_FILL={"Critical":"FBE4E4","Key":"FCEFD6","Non-Key":"E2F4E6"}
SCOPE_FONT={"Critical":"B42318","Key":"9A6700","Non-Key":"1B7A35"}

thin = Side(style="thin", color="D5DBE8")
border_all = Border(left=thin,right=thin,top=thin,bottom=thin)

def fill(hexc): return PatternFill("solid", fgColor=hexc)
def setcell(ws,coord,val,*,bold=False,size=11,color="1F2A44",bg=None,align="left",
            wrap=False,border=False,italic=False,vert="center"):
    c=ws[coord]; c.value=val
    c.font=Font(bold=bold,size=size,color=color,italic=italic,name="Calibri")
    c.alignment=Alignment(horizontal=align,vertical=vert,wrap_text=wrap)
    if bg: c.fill=fill(bg)
    if border: c.border=border_all
    return c

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False
for col,w in {"A":2.5,"B":20,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16,"I":16,"J":16,"K":16,"L":16,"M":16}.items():
    ws.column_dimensions[col].width=w

# Title band
ws.merge_cells("B2:M2")
setcell(ws,"B2","CSA MONTHLY REMINDER  —  EXECUTIVE DASHBOARD",bold=True,size=18,color=WHITE,bg=NAVY,align="left")
ws.row_dimensions[2].height=30
ws.merge_cells("B3:M3")
setcell(ws,"B3","IT General Controls (ITGC) & Access Governance  ·  Source: Archer GRC Search Report  ·  Due 19-Jun-2026  ·  Population: %d control assessments"%total,
        size=10,color=WHITE,bg=BLUE,align="left")
ws.row_dimensions[3].height=18

# KPI cards (row 5-7), each spans 2 cols
kpis=[
 ("CONTROLS PENDING SUBMISSION", total, '100% show CSA Status = "No"', RED),
 ("PAST DUE", "100%", "Due 19-Jun-2026 (elapsed)", RED),
 ("CRITICAL-SCOPE CONTROLS", critical, "%d%% of population"%round(critical/total*100), AMBER),
 ("NO BACKUP DELEGATE", no_dele, "%d%% lack a delegate"%round(no_dele/total*100), AMBER),
 ("FLAGGED 'TO BE RELAUNCHED'", relaunch, "Process / setup backlog", AMBER),
 ("DISTINCT PERFORMERS", performers, "Top owner holds %d"%top_perf[1], SLATE),
]
start_col=2
for i,(label,val,hint,accent) in enumerate(kpis):
    c0=start_col+i*2
    L=get_column_letter(c0); R=get_column_letter(c0+1)
    ws.merge_cells(f"{L}5:{R}5"); ws.merge_cells(f"{L}6:{R}6"); ws.merge_cells(f"{L}7:{R}7")
    setcell(ws,f"{L}5",label,bold=True,size=8.5,color=WHITE,bg=accent,align="center",wrap=True)
    setcell(ws,f"{L}6",val,bold=True,size=26,color=accent,bg=LIGHT,align="center")
    setcell(ws,f"{L}7",hint,size=8.5,color=GREY,bg=LIGHT,align="center",wrap=True)
    for rr in (5,6,7):
        ws[f"{L}{rr}"].border=border_all; ws[f"{R}{rr}"].border=border_all
ws.row_dimensions[5].height=26; ws.row_dimensions[6].height=40; ws.row_dimensions[7].height=26

# ---- helper data tables (placed lower, used as chart sources) ----
def write_table(ws, top, left, title, pairs):
    Lc=get_column_letter(left)
    setcell(ws,f"{Lc}{top}",title,bold=True,size=11,color=NAVY)
    setcell(ws,f"{get_column_letter(left)}{top+1}","Category",bold=True,size=9,color=WHITE,bg=SLATE,border=True)
    setcell(ws,f"{get_column_letter(left+1)}{top+1}","Count",bold=True,size=9,color=WHITE,bg=SLATE,border=True,align="center")
    r=top+2
    for k,v in pairs:
        setcell(ws,f"{get_column_letter(left)}{r}",k,size=10,border=True)
        setcell(ws,f"{get_column_letter(left+1)}{r}",v,size=10,align="center",border=True)
        r+=1
    return top+1, r-1  # header row, last data row

# section header
ws.merge_cells("B9:M9")
setcell(ws,"B9","PORTFOLIO COMPOSITION & OPERATIONAL RISK",bold=True,size=12,color=WHITE,bg=NAVY,align="left")
ws.row_dimensions[9].height=22

# Source tables (put further right / lower so charts overlay nicely). We'll place tables starting row 40.
tbl_row=40
scope_h,scope_e = write_table(ws, tbl_row, 2, "Risk / Control Scope", sorted(by_scope.items(), key=lambda x:-x[1]))
fam_h,fam_e     = write_table(ws, tbl_row, 5, "Control Family", by_family.most_common())
cad_h,cad_e     = write_table(ws, tbl_row, 8, "Review Cadence", by_cadence.most_common())
perf_h,perf_e   = write_table(ws, tbl_row, 11, "Performer Workload", by_perf.most_common())
del_h,del_e     = write_table(ws, tbl_row+ (perf_e-perf_h) + 4, 11, "Delegate Coverage",
                              [("Has delegate", total-no_dele),("No delegate (SPOF)", no_dele)])

# ---- Charts ----
# Scope doughnut
ch1=DoughnutChart(); ch1.title="Risk / Control Scope"; ch1.holeSize=55
data=Reference(ws,min_col=3,min_row=scope_h,max_row=scope_e)
cats=Reference(ws,min_col=2,min_row=scope_h+1,max_row=scope_e)
ch1.add_data(data,titles_from_data=True); ch1.set_categories(cats)
ch1.dataLabels=DataLabelList(); ch1.dataLabels.showVal=True
ch1.height=6.7; ch1.width=8.5
ws.add_chart(ch1,"B10")

# Family bar
ch2=BarChart(); ch2.type="bar"; ch2.title="Control Family"; ch2.legend=None
d2=Reference(ws,min_col=6,min_row=fam_h,max_row=fam_e)
c2=Reference(ws,min_col=5,min_row=fam_h+1,max_row=fam_e)
ch2.add_data(d2,titles_from_data=True); ch2.set_categories(c2)
ch2.dataLabels=DataLabelList(); ch2.dataLabels.showVal=True
ch2.height=6.7; ch2.width=9
ws.add_chart(ch2,"E10")

# Cadence doughnut
ch3=DoughnutChart(); ch3.title="Review Cadence"; ch3.holeSize=55
d3=Reference(ws,min_col=9,min_row=cad_h,max_row=cad_e)
c3=Reference(ws,min_col=8,min_row=cad_h+1,max_row=cad_e)
ch3.add_data(d3,titles_from_data=True); ch3.set_categories(c3)
ch3.dataLabels=DataLabelList(); ch3.dataLabels.showVal=True
ch3.height=6.7; ch3.width=8.5
ws.add_chart(ch3,"I10")

# Performer workload bar
ch4=BarChart(); ch4.type="bar"; ch4.title="Workload by Control Performer"; ch4.legend=None
d4=Reference(ws,min_col=12,min_row=perf_h,max_row=perf_e)
c4=Reference(ws,min_col=11,min_row=perf_h+1,max_row=perf_e)
ch4.add_data(d4,titles_from_data=True); ch4.set_categories(c4)
ch4.dataLabels=DataLabelList(); ch4.dataLabels.showVal=True
ch4.height=9; ch4.width=11
ws.add_chart(ch4,"B25")

# Delegate doughnut
ch5=DoughnutChart(); ch5.title="Backup (Delegate) Coverage"; ch5.holeSize=55
d5=Reference(ws,min_col=12,min_row=del_h,max_row=del_e)
c5=Reference(ws,min_col=11,min_row=del_h+1,max_row=del_e)
ch5.add_data(d5,titles_from_data=True); ch5.set_categories(c5)
ch5.dataLabels=DataLabelList(); ch5.dataLabels.showVal=True
ch5.height=9; ch5.width=9
ws.add_chart(ch5,"I25")

# note under source tables
setcell(ws,"B%d"%(scope_e+3),"Source tables above feed the dashboard charts. Figures derived directly from the Archer export.",italic=True,size=9,color=GREY)

# ============================================================
# SHEET 2: KEY INSIGHTS
# ============================================================
wi = wb.create_sheet("Key Insights")
wi.sheet_view.showGridLines=False
wi.column_dimensions["A"].width=2.5
wi.column_dimensions["B"].width=22
wi.column_dimensions["C"].width=95
wi.merge_cells("B2:C2")
setcell(wi,"B2","KEY INSIGHTS FOR THE EXECUTIVE BOARD",bold=True,size=16,color=WHITE,bg=NAVY)
wi.row_dimensions[2].height=28

insights=[
 ("Compliance · CRITICAL",RED,
  "All %d controls are overdue and unsubmitted"%total,
  'Every assessment shows CSA Submission Status = "No" against a uniform due date of 19-Jun-2026, which has already passed. '
  'Although each control\'s overall workflow shows "Final Closed", the self-assessment attestation step is outstanding across the board — a 100%% non-submission rate.',
  "ACTION: Treat as the highest-priority item this cycle. Escalate to all %d performers and confirm submission within 5 business days."%performers),
 ("Risk Concentration",AMBER,
  "%d of %d controls (%d%%) are Critical scope"%(critical,total,round(critical/total*100)),
  "Critical controls dominate the backlog (Non-Key %d, Key %d). Critical access controls — firefighter/privileged usage, leaver & mover reviews — directly mitigate fraud and SoD risk, so delayed attestation has an outsized control-environment impact."%(by_scope.get('Non-Key',0),by_scope.get('Key',0)),
  "ACTION: Sequence remediation Critical -> Key -> Non-Key. Require sign-off evidence for the %d Critical controls that also have no delegate."%crit_no_dele),
 ("Key-Person Dependency",RED,
  "%d controls (%d%%) have no backup delegate"%(no_dele,round(no_dele/total*100)),
  "More than half the population names only a single performer with no delegate, including %d Critical controls. The heaviest single owner — %s — is responsible for %d controls. This creates execution risk if any individual is absent."%(crit_no_dele,top_perf[0],top_perf[1]),
  "ACTION: Assign named delegates to every Critical and Key control before the next cycle; rebalance the top owner's portfolio."),
 ("Process Hygiene",AMBER,
  "%d controls flagged 'to be relaunched' / 'mail to be sent'"%relaunch,
  'Performer remarks reveal a setup backlog: several controls are pending relaunch ("Relaunch-Monthly", "Annually-To be Relaunched", "Quarterly-To be Relaunched") and a few note open queries ("Ask richard"). This suggests the recurrence schedule in Archer is not fully configured.',
  "ACTION: Have the GRC team finalise recurrence configuration and clear open queries so cadence is enforced automatically."),
 ("Scope Focus",GREEN,
  "Almost entirely Access Governance / ITGC",
  "%d of %d controls sit under Access Governance and %d are ITGC, spanning leaver/mover reviews, privileged & firefighter usage, and access-provisioning approvals. The risk theme is tightly scoped to identity & access management."%(ag,total,itgc),
  "ACTION: A single IAM control owner could coordinate the whole remediation; consider a focused IAM controls war-room."),
 ("Coverage · Traceability",AMBER,
  "Controls span multiple entities & systems",
  "Named systems include Getpaid, L2C, ORIAN, SSIB, VFI, Paloma, IDC Navitro, MOTUS and SAP FACT, across entities such as DE057E, NL153E, NO088L, US006U and the CI* legal entities. %d records carry no entity code, reducing traceability."%no_entity,
  "ACTION: Backfill missing entity codes to improve auditability and ensure each system has an accountable submitter."),
]
r=4
for tag,accent,head,body,rec in insights:
    setcell(wi,f"B{r}",tag.upper(),bold=True,size=9,color=WHITE,bg=accent,align="center",wrap=True,border=True)
    setcell(wi,f"C{r}",head,bold=True,size=12.5,color=NAVY,bg=LIGHT,border=True)
    wi.row_dimensions[r].height=22
    setcell(wi,f"B{r+1}","",bg=LIGHT,border=True)
    setcell(wi,f"C{r+1}",body,size=10.5,color=SLATE,wrap=True,border=True,vert="top")
    wi.row_dimensions[r+1].height=58
    setcell(wi,f"B{r+2}","",bg="F4F7FE",border=True)
    setcell(wi,f"C{r+2}",rec,bold=True,size=10,color="1B4DB1",bg="F4F7FE",wrap=True,border=True,vert="top")
    wi.row_dimensions[r+2].height=30
    r+=4

# recommended sequencing
setcell(wi,f"B{r}","RECOMMENDED SEQUENCING",bold=True,size=11,color=WHITE,bg=NAVY)
wi.merge_cells(f"B{r}:C{r}")
seq="1) Escalate all %d submissions immediately  ->  2) Remediate Critical -> Key -> Non-Key  ->  3) Assign delegates to every Critical/Key control  ->  4) GRC to finalise recurrence configuration."%total
setcell(wi,f"B{r+1}",seq,size=10.5,color=SLATE,wrap=True,bg=LIGHT,border=True)
wi.merge_cells(f"B{r+1}:C{r+1}")
wi.row_dimensions[r+1].height=34

# ============================================================
# SHEET 3: CONTROL REGISTER (detail data)
# ============================================================
wd = wb.create_sheet("Control Register")
wd.sheet_view.showGridLines=False
cols=[("Assessment ID","Assessment ID",14),("Control Procedure","Control Procedure",26),
      ("Control Procedure Name","Control Name",55),("Risk/Control Scope","Scope",12),
      ("Control Performer","Performer",26),("Control Performer Delegate","Delegate",26),
      ("Entity Code","Entity",10),("Business Process","Business Process",20),
      ("Business Sub Process","Sub Process",22),("Due Date","Due Date",12),
      ("CSA Submission Status","CSA Submitted",13),("Remarks","Remarks",26),("Tracking ID","Tracking ID",13)]
# title
wd.merge_cells("A1:M1")
setcell(wd,"A1","CONTROL REGISTER — CSA Monthly Reminder (%d controls)"%total,bold=True,size=14,color=WHITE,bg=NAVY)
wd.row_dimensions[1].height=24
# header
hr=2
for j,(src,disp,w) in enumerate(cols,start=1):
    L=get_column_letter(j); wd.column_dimensions[L].width=w
    setcell(wd,f"{L}{hr}",disp,bold=True,size=9.5,color=WHITE,bg=SLATE,align="center",wrap=True,border=True)
wd.row_dimensions[hr].height=28
# sort by scope severity then performer
order={"Critical":0,"Key":1,"Non-Key":2}
srecs=sorted(records,key=lambda d:(order.get(d.get("Risk/Control Scope"),9), d.get("Control Performer") or ""))
rr=hr+1
for d in srecs:
    sc=d.get("Risk/Control Scope") or ""
    for j,(src,disp,w) in enumerate(cols,start=1):
        L=get_column_letter(j)
        val=d.get(src)
        if src=="Control Performer Delegate" and not val: val="— none —"
        if val is None: val=""
        c=setcell(wd,f"{L}{rr}",val,size=9.5,border=True,vert="top",
                  wrap=(src in("Control Procedure Name","Remarks")))
        if src=="Risk/Control Scope" and sc in SCOPE_FILL:
            c.fill=fill(SCOPE_FILL[sc]); c.font=Font(bold=True,size=9.5,color=SCOPE_FONT[sc]); c.alignment=Alignment(horizontal="center",vertical="center")
        if src=="CSA Submission Status":
            c.fill=fill("FBE4E4"); c.font=Font(bold=True,size=9.5,color="B42318"); c.alignment=Alignment(horizontal="center",vertical="center")
        if src=="Control Performer Delegate" and val=="— none —":
            c.font=Font(italic=True,size=9.5,color="B42318")
    wd.row_dimensions[rr].height=30
    rr+=1
wd.freeze_panes="A3"
wd.auto_filter.ref=f"A{hr}:M{rr-1}"

# ============================================================
# SHEET 4: SUMMARY TABLES (clean pivots)
# ============================================================
wsm=wb.create_sheet("Summary Tables")
wsm.sheet_view.showGridLines=False
wsm.column_dimensions["A"].width=2.5
for c in "BCDE": wsm.column_dimensions[c].width=30
setcell(wsm,"B2","SUMMARY / PIVOT TABLES",bold=True,size=14,color=WHITE,bg=NAVY); wsm.merge_cells("B2:E2")
wsm.row_dimensions[2].height=24
def pivot(ws,top,title,pairs,pct_base=None):
    setcell(ws,f"B{top}",title,bold=True,size=11,color=NAVY)
    setcell(ws,f"B{top+1}","Category",bold=True,color=WHITE,bg=SLATE,border=True)
    setcell(ws,f"C{top+1}","Count",bold=True,color=WHITE,bg=SLATE,border=True,align="center")
    setcell(ws,f"D{top+1}","% of total",bold=True,color=WHITE,bg=SLATE,border=True,align="center")
    r=top+2; tot=sum(v for _,v in pairs)
    for k,v in pairs:
        setcell(ws,f"B{r}",k,border=True)
        setcell(ws,f"C{r}",v,align="center",border=True)
        setcell(ws,f"D{r}","%d%%"%round(v/tot*100),align="center",border=True)
        r+=1
    setcell(ws,f"B{r}","Total",bold=True,bg=LIGHT,border=True)
    setcell(ws,f"C{r}",tot,bold=True,align="center",bg=LIGHT,border=True)
    setcell(ws,f"D{r}","100%",bold=True,align="center",bg=LIGHT,border=True)
    return r+3
nxt=4
nxt=pivot(wsm,nxt,"By Risk / Control Scope",sorted(by_scope.items(),key=lambda x:-x[1]))
nxt=pivot(wsm,nxt,"By Control Family",by_family.most_common())
nxt=pivot(wsm,nxt,"By Review Cadence",by_cadence.most_common())
nxt=pivot(wsm,nxt,"By Control Performer",by_perf.most_common())
nxt=pivot(wsm,nxt,"Delegate Coverage",[("Has delegate",total-no_dele),("No delegate",no_dele)])
nxt=pivot(wsm,nxt,"By Business Process",cnt(lambda d:d.get("Business Process")).most_common())

wb.active=0
wb.save(OUT)
print("Saved", OUT, "| sheets:", wb.sheetnames)
